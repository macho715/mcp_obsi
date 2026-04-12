import json
from pathlib import Path

from openpyxl import Workbook
from rdflib import Graph, Literal, Namespace, URIRef

import scripts.build_dashboard_graph_data as dashboard_export
from app.services.graph_projection_builder import build_dashboard_projection
from scripts.build_dashboard_graph_data import export_dashboard_graph_data


def _write_status_excel(excel_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(
        [
            "SCT SHIP NO.",
            "PO No.",
            "VENDOR",
            "VESSEL NAME/ FLIGHT No.",
            "MOSB",
            "DSV Indoor",
            "AGI",
        ]
    )
    worksheet.append(
        [
            "SHIP-001",
            "PO-001",
            "Vendor A",
            "Vessel A",
            "2026-04-01",
            "2026-04-02",
            "2026-04-03",
        ]
    )
    workbook.save(excel_path)


def _write_analysis_note(path: Path, title: str, slug: str) -> None:
    path.write_text(
        f"""---
title: {title}
slug: {slug}
tags:
  - agi
  - mosb
---
Issue body
""",
        encoding="utf-8",
    )


def _write_route_status_excel(excel_path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet.append(
        [
            "SCT SHIP NO.",
            "PO No.",
            "VENDOR",
            "VESSEL NAME/ FLIGHT No.",
            "COE",
            "POL",
            "POD",
            "SHIP MODE",
            "ATD",
            "ATA",
        ]
    )
    worksheet.append(
        [
            "SHIP-001",
            "PO-001",
            "Vendor A",
            "Vessel A",
            "FRANCE",
            "Le Havre",
            "Mina Zayed",
            "Sea Freight",
            "2023-11-12",
            "2023-12-01",
        ]
    )
    workbook.save(excel_path)


def test_build_dashboard_projection_emits_flat_consumer_contract():
    shipment_id = "https://hvdc.logistics/resource/shipment/HVDC-001"
    event_id = "https://hvdc.logistics/resource/arrival/HVDC-001/agi/2026-04-01"
    location_id = "https://hvdc.logistics/resource/site/agi"
    lesson_id = "https://hvdc.logistics/resource/lesson/HVDC-001/lesson-1"

    nodes, edges, audits = build_dashboard_projection(
        shipments=[
            {
                "id": shipment_id,
                "label": "HVDC-001",
                "type": "Shipment",
            }
        ],
        events=[
            {
                "id": event_id,
                "label": "AGI",
                "type": "ArrivalEvent",
                "subject_id": shipment_id,
                "location_id": location_id,
            }
        ],
        lessons=[
            {
                "id": lesson_id,
                "label": "Delay at AGI",
                "type": "IncidentLesson",
                "shipment_id": shipment_id,
                "location_id": location_id,
            }
        ],
    )

    node_by_id = {node["data"]["id"]: node["data"] for node in nodes}
    edge_keys = {
        (edge["data"]["source"], edge["data"]["target"], edge["data"]["label"]) for edge in edges
    }

    assert shipment_id in node_by_id
    assert node_by_id[shipment_id]["type"] == "Shipment"
    assert location_id in node_by_id
    assert node_by_id[location_id]["type"] == "SiteLocation"
    assert (shipment_id, location_id, "occurredAt") in edge_keys
    assert (shipment_id, lesson_id, "relatedToLesson") in edge_keys
    assert audits["projection"]["unknown_nodes"] == 0


def test_export_dashboard_graph_data_wires_route_leg_and_milestone_data_into_ttl(
    tmp_path: Path,
):
    excel_path = tmp_path / "hvdc_status.xlsx"
    output_dir = tmp_path / "out"
    ttl_path = tmp_path / "knowledge_graph.ttl"

    _write_route_status_excel(excel_path)

    export_dashboard_graph_data(
        excel_path=excel_path,
        wiki_dir=tmp_path / "missing-analyses",
        output_dir=output_dir,
        ttl_path=ttl_path,
    )

    graph = Graph()
    graph.parse(ttl_path, format="turtle")
    hvdc = Namespace("http://hvdc.logistics/ontology/")

    shipment_id = URIRef("https://hvdc.logistics/resource/shipment/SHIP-001")
    leg_id = URIRef("https://hvdc.logistics/resource/journey-leg/SHIP-001/main")
    pol_id = URIRef("https://hvdc.logistics/resource/port/le_havre")
    pod_id = URIRef("https://hvdc.logistics/resource/port/mina_zayed")
    atd_id = URIRef("https://hvdc.logistics/resource/milestone/SHIP-001/M61")
    ata_id = URIRef("https://hvdc.logistics/resource/milestone/SHIP-001/M80")

    assert (shipment_id, hvdc.countryOfExport, Literal("FRANCE")) in graph
    assert (shipment_id, hvdc.portOfLoading, Literal("Le Havre")) in graph
    assert (shipment_id, hvdc.portOfDischarge, Literal("Mina Zayed")) in graph
    assert (shipment_id, hvdc.shipMode, Literal("SEA")) in graph
    assert (shipment_id, hvdc.hasJourneyLeg, leg_id) in graph
    assert (leg_id, hvdc.originPort, pol_id) in graph
    assert (leg_id, hvdc.destinationPort, pod_id) in graph
    assert (shipment_id, hvdc.hasMilestone, atd_id) in graph
    assert (shipment_id, hvdc.hasMilestone, ata_id) in graph
    assert (atd_id, hvdc.milestoneCode, Literal("M61")) in graph
    assert (ata_id, hvdc.milestoneCode, Literal("M80")) in graph


def test_export_dashboard_graph_data_exposes_route_and_timing_mirrors_on_shipment_node(
    tmp_path: Path,
):
    excel_path = tmp_path / "hvdc_status.xlsx"
    output_dir = tmp_path / "out"

    _write_route_status_excel(excel_path)

    export_dashboard_graph_data(
        excel_path=excel_path,
        wiki_dir=tmp_path / "missing-analyses",
        output_dir=output_dir,
        ttl_path=None,
    )

    nodes = json.loads((output_dir / "nodes.json").read_text(encoding="utf-8"))
    shipment = next(node["data"] for node in nodes if node["data"]["type"] == "Shipment")

    assert shipment["countryOfExport"] == "FRANCE"
    assert shipment["portOfLoading"] == "Le Havre"
    assert shipment["portOfDischarge"] == "Mina Zayed"
    assert shipment["shipMode"] == "SEA"
    assert shipment["actualDeparture"] == "2023-11-12"
    assert shipment["actualArrival"] == "2023-12-01"


def test_export_dashboard_graph_data_emits_consumer_contract(tmp_path: Path):
    excel_path = tmp_path / "hvdc_status.xlsx"
    wiki_root = tmp_path / "contract-vault"
    wiki_dir = wiki_root / "wiki" / "analyses"
    wiki_dir.mkdir(parents=True)
    (wiki_root / ".obsidian").mkdir()
    output_dir = tmp_path / "out"
    audit_dir = tmp_path / "audits"

    _write_status_excel(excel_path)
    _write_analysis_note(
        wiki_dir / "logistics_issue_delay_at_agi.md", "Delay at AGI", "delay-at-agi"
    )

    export_dashboard_graph_data(
        excel_path=excel_path,
        wiki_dir=wiki_dir,
        output_dir=output_dir,
        audit_dir=audit_dir,
    )

    nodes_path = output_dir / "nodes.json"
    edges_path = output_dir / "edges.json"
    nodes = json.loads(nodes_path.read_text(encoding="utf-8"))
    edges = json.loads(edges_path.read_text(encoding="utf-8"))

    export_dashboard_graph_data(
        excel_path=excel_path,
        wiki_dir=wiki_dir,
        output_dir=output_dir,
    )
    assert json.loads(nodes_path.read_text(encoding="utf-8")) == nodes
    assert json.loads(edges_path.read_text(encoding="utf-8")) == edges

    assert nodes
    assert edges
    assert all("data" in node for node in nodes)
    required_node_keys = {"id", "label", "type", "rdf-schema#label"}
    required_edge_keys = {"id", "source", "target", "label"}
    assert all(required_node_keys <= set(node["data"].keys()) for node in nodes)
    assert all("data" in edge for edge in edges)
    assert all(required_edge_keys <= set(edge["data"].keys()) for edge in edges)

    node_ids = {node["data"]["id"] for node in nodes}
    assert all(edge["data"]["source"] in node_ids for edge in edges)
    assert all(edge["data"]["target"] in node_ids for edge in edges)
    assert [node["data"]["id"] for node in nodes] == [
        node["data"]["id"]
        for node in sorted(
            nodes,
            key=lambda node: (node["data"]["type"], node["data"]["label"], node["data"]["id"]),
        )
    ]
    assert [edge["data"]["id"] for edge in edges] == [
        edge["data"]["id"]
        for edge in sorted(
            edges,
            key=lambda edge: (
                edge["data"]["source"],
                edge["data"]["target"],
                edge["data"]["label"],
            ),
        )
    ]

    issue_nodes = [node for node in nodes if node["data"]["type"] == "LogisticsIssue"]
    lesson_nodes = [node for node in nodes if node["data"]["type"] == "IncidentLesson"]
    assert len(issue_nodes) == 1
    assert issue_nodes[0]["data"]["label"] == "Delay at AGI"
    assert issue_nodes[0]["data"]["analysisPath"] == "wiki/analyses/logistics_issue_delay_at_agi.md"
    assert issue_nodes[0]["data"]["analysisVault"] == "contract-vault"
    assert len(lesson_nodes) == 1
    assert (
        lesson_nodes[0]["data"]["analysisPath"] == "wiki/analyses/logistics_issue_delay_at_agi.md"
    )
    assert lesson_nodes[0]["data"]["analysisVault"] == "contract-vault"
    assert all(node["data"]["type"] != "rdf-schema#Class" for node in nodes)
    assert all(node["data"]["type"] != "Unknown" for node in nodes)

    node_ids = {node["data"]["id"] for node in nodes}
    shipment_legacy_id = "http://hvdc.logistics/ontology/shipment/SHIP-001"
    shipment_canonical_id = "https://hvdc.logistics/resource/shipment/SHIP-001"
    location_legacy_id = "http://hvdc.logistics/ontology/site/AGI"
    location_canonical_id = "https://hvdc.logistics/resource/site/agi"
    lesson_id = "http://hvdc.logistics/ontology/lesson/delay-at-agi"

    assert shipment_legacy_id in node_ids
    assert shipment_canonical_id not in node_ids
    assert location_legacy_id in node_ids
    assert location_canonical_id not in node_ids
    assert (
        sum(
            1
            for node in nodes
            if node["data"]["type"] == "Shipment" and node["data"]["label"] == "SHIP-001"
        )
        == 1
    )
    assert (location_legacy_id, lesson_id, "relatedToLesson") in {
        (edge["data"]["source"], edge["data"]["target"], edge["data"]["label"]) for edge in edges
    }

    source_audit = json.loads(
        (audit_dir / "hvdc_ttl_source_audit.json").read_text(encoding="utf-8")
    )
    assert source_audit["selected_analyses_dir"] == str(wiki_dir)
    assert source_audit["analyses_dir_fallback_used"] is False


def test_export_dashboard_graph_data_prefers_external_analyses_dir_when_available(
    tmp_path: Path, monkeypatch
):
    excel_path = tmp_path / "hvdc_status.xlsx"
    external_vault = tmp_path / "external-vault"
    fallback_vault = tmp_path / "vault"
    external_analyses = external_vault / "wiki" / "analyses"
    fallback_analyses = fallback_vault / "wiki" / "analyses"
    external_analyses.mkdir(parents=True)
    fallback_analyses.mkdir(parents=True)
    (external_vault / ".obsidian").mkdir()
    (fallback_vault / ".obsidian").mkdir()
    output_dir = tmp_path / "out"
    audit_dir = tmp_path / "audits"

    _write_status_excel(excel_path)
    _write_analysis_note(
        external_analyses / "logistics_issue_delay_at_agi.md",
        "Delay at AGI",
        "delay-at-agi",
    )
    _write_analysis_note(
        fallback_analyses / "logistics_issue_fallback.md",
        "Fallback Note",
        "fallback-note",
    )

    monkeypatch.setattr(dashboard_export, "PRIMARY_ANALYSES_DIR", external_analyses)
    monkeypatch.setattr(dashboard_export, "DEFAULT_WIKI_DIR", fallback_analyses)

    export_dashboard_graph_data(
        excel_path=excel_path,
        wiki_dir=None,
        output_dir=output_dir,
        audit_dir=audit_dir,
        ttl_path=None,
    )

    nodes = json.loads((output_dir / "nodes.json").read_text(encoding="utf-8"))
    issue_node = next(node["data"] for node in nodes if node["data"]["type"] == "LogisticsIssue")
    lesson_node = next(node["data"] for node in nodes if node["data"]["type"] == "IncidentLesson")
    source_audit = json.loads(
        (audit_dir / "hvdc_ttl_source_audit.json").read_text(encoding="utf-8")
    )

    assert source_audit["selected_analyses_dir"] == str(external_analyses.resolve())
    assert source_audit["analyses_dir_fallback_used"] is False
    assert issue_node["analysisPath"] == "wiki/analyses/logistics_issue_delay_at_agi.md"
    assert issue_node["analysisVault"] == "external-vault"
    assert lesson_node["analysisPath"] == "wiki/analyses/logistics_issue_delay_at_agi.md"
    assert lesson_node["analysisVault"] == "external-vault"


def test_export_dashboard_graph_data_falls_back_to_repo_local_analyses_dir_when_external_missing(
    tmp_path: Path, monkeypatch
):
    excel_path = tmp_path / "hvdc_status.xlsx"
    fallback_vault = tmp_path / "vault"
    fallback_analyses = fallback_vault / "wiki" / "analyses"
    fallback_analyses.mkdir(parents=True)
    (fallback_vault / ".obsidian").mkdir()
    output_dir = tmp_path / "out"
    audit_dir = tmp_path / "audits"

    _write_status_excel(excel_path)
    _write_analysis_note(
        fallback_analyses / "logistics_issue_delay_at_agi.md",
        "Delay at AGI",
        "delay-at-agi",
    )

    monkeypatch.setattr(dashboard_export, "PRIMARY_ANALYSES_DIR", tmp_path / "missing-external")
    monkeypatch.setattr(dashboard_export, "DEFAULT_WIKI_DIR", fallback_analyses)

    export_dashboard_graph_data(
        excel_path=excel_path,
        wiki_dir=None,
        output_dir=output_dir,
        audit_dir=audit_dir,
        ttl_path=None,
    )

    nodes = json.loads((output_dir / "nodes.json").read_text(encoding="utf-8"))
    issue_node = next(node["data"] for node in nodes if node["data"]["type"] == "LogisticsIssue")
    lesson_node = next(node["data"] for node in nodes if node["data"]["type"] == "IncidentLesson")
    source_audit = json.loads(
        (audit_dir / "hvdc_ttl_source_audit.json").read_text(encoding="utf-8")
    )

    assert source_audit["selected_analyses_dir"] == str(fallback_analyses.resolve())
    assert source_audit["analyses_dir_fallback_used"] is True
    assert issue_node["analysisPath"] == "wiki/analyses/logistics_issue_delay_at_agi.md"
    assert issue_node["analysisVault"] == "vault"
    assert lesson_node["analysisPath"] == "wiki/analyses/logistics_issue_delay_at_agi.md"
    assert lesson_node["analysisVault"] == "vault"


def test_export_dashboard_graph_data_keeps_legacy_id_style_for_values_with_spaces_and_slashes(
    tmp_path: Path,
):
    excel_path = tmp_path / "hvdc_status.xlsx"
    wiki_dir = tmp_path / "wiki" / "analyses"
    wiki_dir.mkdir(parents=True)
    (tmp_path / ".obsidian").mkdir()
    output_dir = tmp_path / "out"

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["SCT SHIP NO.", "PO No.", "VENDOR", "VESSEL NAME/ FLIGHT No."])
    worksheet.append(["SHIP 002", "PO 002", "BMT Co., Ltd", "EK0118 / EK09113"])
    workbook.save(excel_path)

    export_dashboard_graph_data(
        excel_path=excel_path,
        wiki_dir=wiki_dir,
        output_dir=output_dir,
    )

    nodes = json.loads((output_dir / "nodes.json").read_text(encoding="utf-8"))
    edges = json.loads((output_dir / "edges.json").read_text(encoding="utf-8"))
    node_ids = {node["data"]["id"] for node in nodes}
    required_edge_keys = {"id", "source", "target", "label"}
    assert all(required_edge_keys <= set(edge["data"].keys()) for edge in edges)
    assert all(edge["data"]["source"] in node_ids for edge in edges)
    assert all(edge["data"]["target"] in node_ids for edge in edges)

    assert "http://hvdc.logistics/ontology/vendor/BMT_Co.,_Ltd" in node_ids
    assert "http://hvdc.logistics/ontology/vessel/EK0118_/_EK09113" in node_ids
