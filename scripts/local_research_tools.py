from __future__ import annotations

from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from scripts.local_wiki_everything import search_everything
from scripts.local_wiki_extract import extract_document

DENIED_PATH_MARKERS = (
    "\\.git\\",
    "\\.venv\\",
    "\\node_modules\\",
    "\\.codex\\",
    "\\.cursor\\",
)
CREDENTIAL_PATH_MARKERS = (
    "password",
    "secret",
    "token",
    "apikey",
    "api-key",
    "private-key",
    "credential",
    "credentials",
)

SearchFn = Callable[[str, int], list[dict[str, Any]]]
ExtractFn = Callable[[Path], dict[str, Any]]


@dataclass(frozen=True)
class ToolLimits:
    max_rounds: int = 5
    max_search_calls: int = 5
    max_search_results: int = 20
    max_extract_files: int = 10
    max_chars_per_file: int = 12_000
    max_total_chars: int = 80_000
    max_table_preview_sheets: int = 5
    max_table_preview_rows: int = 20
    max_table_preview_columns: int = 12
    max_compare_files: int = 4


class LocalResearchToolExecutor:
    def __init__(
        self,
        *,
        selected_candidates: list[dict[str, Any]] | None = None,
        preview_candidates: list[dict[str, Any]] | None = None,
        search: SearchFn | None = None,
        extract: ExtractFn | None = None,
        limits: ToolLimits | None = None,
    ) -> None:
        self.search = search or search_everything
        self.extract = extract or extract_document
        self.limits = limits or ToolLimits()
        self.search_calls = 0
        self.extracted_files = 0
        self.total_chars = 0
        self.allowed_paths: set[str] = set()
        self.candidate_metadata: dict[str, dict[str, Any]] = {}
        self._add_candidates(selected_candidates or [])
        self._add_candidates(preview_candidates or [])

    def execute(self, tool_name: str, arguments: dict[str, Any] | None = None) -> dict[str, Any]:
        args = arguments or {}
        if tool_name == "everything_search":
            return self.everything_search(
                str(args.get("query") or ""), limit=_as_int(args.get("limit"), 20)
            )
        if tool_name == "extract_file":
            return self.extract_file(str(args.get("path") or ""))
        if tool_name == "extract_table_preview":
            return self.extract_table_preview(str(args.get("path") or ""))
        if tool_name == "compare_selected_files":
            paths = args.get("paths")
            if not isinstance(paths, list):
                paths = []
            return self.compare_selected_files([str(path) for path in paths])
        if tool_name == "build_citation":
            return self.build_citation(
                str(args.get("path") or ""),
                excerpt=str(args.get("excerpt") or args.get("evidence") or ""),
            )
        return _rejected("unknown tool", tool=tool_name)

    def everything_search(self, query: str, *, limit: int | None = None) -> dict[str, Any]:
        if not query.strip():
            return _rejected("query must not be empty", tool="everything_search")
        if self.search_calls >= self.limits.max_search_calls:
            return _rejected("search call limit reached", tool="everything_search")

        self.search_calls += 1
        effective_limit = max(
            1, min(_as_int(limit, self.limits.max_search_results), self.limits.max_search_results)
        )
        try:
            raw_results = self.search(query, effective_limit)
        except Exception as exc:
            return _rejected(f"{type(exc).__name__}: {exc}", tool="everything_search")

        results: list[dict[str, Any]] = []
        rejected: list[dict[str, Any]] = []
        for item in raw_results[:effective_limit]:
            payload = _candidate_payload(item)
            guard = self._path_guard(payload["path"])
            if guard is not None:
                rejected.append(guard)
                continue
            self._allow_candidate(payload)
            results.append(payload)

        return {
            "ok": True,
            "status": "ok",
            "tool": "everything_search",
            "query": query,
            "limit": effective_limit,
            "results": results,
            "rejected": rejected,
        }

    def extract_file(self, path: str) -> dict[str, Any]:
        guard = self._authorization_guard(path)
        if guard is not None:
            return guard
        if self.extracted_files >= self.limits.max_extract_files:
            return _rejected("extract file limit reached", tool="extract_file", path=path)

        source = Path(path)
        try:
            extracted = self.extract(source)
        except Exception as exc:
            return _rejected(f"{type(exc).__name__}: {exc}", tool="extract_file", path=path)

        self.extracted_files += 1
        text = str(extracted.get("text") or "")
        capped, truncated = self._cap_text(text)
        return {
            "ok": True,
            "tool": "extract_file",
            "path": str(source),
            "name": source.name,
            "extension": source.suffix.lower(),
            "status": str(extracted.get("status") or "ok"),
            "reason": str(extracted.get("reason") or ""),
            "text": capped,
            "truncated": truncated,
        }

    def extract_table_preview(self, path: str) -> dict[str, Any]:
        guard = self._authorization_guard(path)
        if guard is not None:
            return guard

        source = Path(path)
        if source.suffix.lower() not in {".xlsx", ".xlsm"}:
            return _rejected(
                "table preview supports .xlsx and .xlsm", tool="extract_table_preview", path=path
            )

        try:
            workbook = load_workbook(source, read_only=True, data_only=True)
        except Exception as exc:
            return _rejected(
                f"{type(exc).__name__}: {exc}", tool="extract_table_preview", path=path
            )

        sheets: list[dict[str, Any]] = []
        sheet_count = 0
        try:
            sheet_count = len(workbook.sheetnames)
            for sheet_index, worksheet in enumerate(workbook.worksheets):
                if sheet_index >= self.limits.max_table_preview_sheets:
                    break
                rows: list[list[str]] = []
                for row_index, row in enumerate(
                    worksheet.iter_rows(
                        min_row=1,
                        max_row=self.limits.max_table_preview_rows,
                        max_col=self.limits.max_table_preview_columns,
                        values_only=True,
                    )
                ):
                    if row_index >= self.limits.max_table_preview_rows:
                        break
                    values = ["" if value is None else str(value) for value in row]
                    if any(values):
                        rows.append(values)
                sheets.append({"name": worksheet.title, "rows": rows})
        finally:
            workbook.close()

        return {
            "ok": True,
            "status": "ok",
            "tool": "extract_table_preview",
            "path": str(source),
            "name": source.name,
            "extension": source.suffix.lower(),
            "sheets": sheets,
            "truncated": sheet_count > self.limits.max_table_preview_sheets,
        }

    def compare_selected_files(self, paths: list[str]) -> dict[str, Any]:
        files: list[dict[str, Any]] = []
        rejected: list[dict[str, Any]] = []
        for path in paths[: self.limits.max_compare_files]:
            extracted = self.extract_file(path)
            if extracted.get("status") == "ok":
                files.append(
                    {
                        "path": extracted["path"],
                        "name": extracted["name"],
                        "extension": extracted["extension"],
                        "status": extracted["status"],
                        "text": extracted["text"],
                    }
                )
            else:
                rejected.append(extracted)

        return {
            "ok": True,
            "status": "ok",
            "tool": "compare_selected_files",
            "files": files,
            "rejected": rejected,
            "truncated": len(paths) > self.limits.max_compare_files,
        }

    def build_citation(self, path: str, excerpt: str = "", *, label: str = "") -> dict[str, Any]:
        _ = label
        return {"source_path": path, "evidence": excerpt}

    def _add_candidates(self, candidates: list[dict[str, Any]]) -> None:
        for candidate in candidates:
            payload = _candidate_payload(candidate)
            if self._path_guard(payload["path"]) is None:
                self._allow_candidate(payload)

    def _allow_candidate(self, candidate: dict[str, Any]) -> None:
        key = _path_key(candidate["path"])
        self.allowed_paths.add(key)
        self.candidate_metadata[key] = candidate

    def _authorization_guard(self, path: str) -> dict[str, Any] | None:
        path_guard = self._path_guard(path)
        if path_guard is not None:
            return path_guard
        if _path_key(path) not in self.allowed_paths:
            return _rejected(
                "path not approved by selected, preview, or same-request search",
                tool="path_guard",
                path=path,
            )
        return None

    def _path_guard(self, path: str) -> dict[str, Any] | None:
        if not path.strip():
            return _rejected("path must not be empty", tool="path_guard", path=path)
        normalized = _normalized_path_text(path)
        if any(marker in normalized for marker in DENIED_PATH_MARKERS):
            return _rejected("blocked path: denied workspace", tool="path_guard", path=path)
        if any(marker in normalized for marker in CREDENTIAL_PATH_MARKERS):
            return _rejected("blocked path: credential-looking path", tool="path_guard", path=path)
        return None

    def _cap_text(self, text: str) -> tuple[str, bool]:
        remaining = max(0, self.limits.max_total_chars - self.total_chars)
        limit = min(self.limits.max_chars_per_file, remaining)
        capped = text[:limit]
        self.total_chars += len(capped)
        return capped, len(capped) < len(text)


def _candidate_payload(candidate: dict[str, Any]) -> dict[str, Any]:
    path = str(candidate.get("path") or "").strip()
    source = Path(path)
    name = str(candidate.get("name") or source.name).strip()
    extension = str(candidate.get("extension") or source.suffix).lower()
    payload = {
        "path": path,
        "name": name,
        "extension": extension,
    }
    for key in ("size", "modifiedAt", "modified_at", "score", "rank_reason", "status"):
        if key in candidate:
            payload[key] = candidate[key]
    return payload


def _rejected(reason: str, *, tool: str, path: str = "") -> dict[str, Any]:
    payload: dict[str, Any] = {
        "ok": False,
        "status": "rejected",
        "tool": tool,
        "reason": reason,
        "error": {
            "code": "rejected",
            "message": reason,
        },
    }
    if path:
        payload["path"] = path
    return payload


def _path_key(path: str) -> str:
    return str(Path(path)).replace("/", "\\").casefold()


def _normalized_path_text(path: str) -> str:
    normalized = _path_key(path)
    if not normalized.startswith("\\"):
        normalized = "\\" + normalized
    if not normalized.endswith("\\"):
        normalized += "\\"
    return normalized


def _as_int(value: Any, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default
