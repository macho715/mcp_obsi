from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook
from rdflib import Graph, Namespace, URIRef

from app.services.graph_canonical_builder import build_canonical_graph
from app.services.graph_knowledge_builder import build_knowledge_objects
from app.services.graph_mapping_builder import build_compatibility_mappings
from app.services.graph_normalizer import normalize_sources as normalize_graph_sources
from app.services.graph_projection_builder import build_dashboard_projection
from app.services.graph_resolver import resolve_analysis_note, resolve_location
from app.services.graph_source_loader import REQUIRED_JPT_SHEETS
from app.services.graph_types import CanonicalCase, CanonicalEvent, CanonicalShipment

HVDC_BASE = "http://hvdc.logistics/ontology/"
HVDC = Namespace(HVDC_BASE)

DEFAULT_EXCEL_PATH = Path("Logi ontol core doc/HVDC STATUS.xlsx")
DEFAULT_WAREHOUSE_STATUS_PATH = Path("Logi ontol core doc/HVDC WAREHOUSE STATUS.xlsx")
DEFAULT_JPT_RECONCILED_PATH = Path("Logi ontol core doc/JPT-reconciled_v6.0.xlsx")
DEFAULT_INLAND_COST_PATH = Path("Logi ontol core doc/HVDC Logistics cost(inland,domestic).xlsx")
DEFAULT_WIKI_DIR = Path("vault/wiki/analyses")
PRIMARY_ANALYSES_DIR = Path(r"C:\Users\jichu\Downloads\valut\wiki\analyses")
DEFAULT_OUTPUT_DIR = Path("kg-dashboard/public/data")
DEFAULT_TTL_PATH = Path("vault/knowledge_graph.ttl")
DEFAULT_AUDIT_DIR = Path("runtime/audits")

REQUIRED_EXCEL_COLUMNS = [
    "SCT SHIP NO.",
    "PO No.",
    "VENDOR",
    "VESSEL NAME/ FLIGHT No.",
]

WAREHOUSE_COLUMNS = [
    "DSV Indoor",
    "DSV Outdoor",
    "DSV MZD",
    "DSV Kizad",
    "JDN MZD",
    "JDN Waterfront",
    "AAA Storage",
    "ZENER (WH)",
    "Hauler DG Storage",
    "Vijay Tanks",
]

SITE_COLUMNS = ["SHU", "MIR", "DAS", "AGI"]


@dataclass(slots=True)
class LoadedGraphSources:
    shipment_rows: list[dict[str, Any]] = field(default_factory=list)
    warehouse_rows: list[dict[str, Any]] = field(default_factory=list)
    jpt_sheets: dict[str, list[dict[str, Any]]] = field(default_factory=dict)
    inland_cost_rows: list[dict[str, Any]] = field(default_factory=list)
    analysis_notes: list[dict[str, Any]] = field(default_factory=list)


@dataclass(slots=True)
class NormalizedSources:
    shipments: list[CanonicalShipment]
    cases: list[CanonicalCase]
    route_events: list[CanonicalEvent]
    document_refs: list[dict[str, Any]]
    status_snapshots: list[dict[str, Any]]
    charge_candidates: list[dict[str, Any]]


def normalize_fragment(value: str) -> str:
    return re.sub(r"\s+", "_", value.strip())


def node_uri(kind: str, value: str) -> str:
    return str(URIRef(HVDC[f"{kind}/{normalize_fragment(value)}"]))


def _safe_uri(value: object | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    return text.replace(" ", "_")


def _legacy_dashboard_id(value: object | None) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if text.startswith("http://hvdc.logistics/ontology/"):
        return text

    for prefix in (
        "https://hvdc.logistics/resource/",
        "http://hvdc.logistics/resource/",
    ):
        if not text.startswith(prefix):
            continue
        parts = text.rstrip("/").split("/")
        if len(parts) < 2:
            return text
        kind = parts[-2]
        value_text = parts[-1]
        if kind in {"site", "hub"}:
            value_text = value_text.upper()
        if kind in {"shipment", "site", "hub", "carrier", "pattern"}:
            return node_uri(kind, value_text)
        return text

    return text


def _is_present(value: object) -> bool:
    if value is None:
        return False
    text = str(value).strip()
    return bool(text) and text.lower() not in {"nan", "none"}


def _read_frontmatter(path: Path) -> dict[str, Any]:
    content = path.read_text(encoding="utf-8")
    if not content.startswith("---"):
        return {}
    parts = content.split("---", 2)
    if len(parts) < 3:
        return {}
    try:
        frontmatter = yaml.safe_load(parts[1]) or {}
    except yaml.YAMLError:
        return {}
    return frontmatter if isinstance(frontmatter, dict) else {}


def _find_obsidian_vault_root(path: Path | None) -> Path | None:
    if path is None:
        return None

    current = path.resolve()
    while True:
        if (current / ".obsidian").exists():
            return current
        if current.parent == current:
            return None
        current = current.parent


def _analysis_note_metadata(note_path: Path, analyses_dir: Path | None) -> dict[str, str | None]:
    vault_root = _find_obsidian_vault_root(analyses_dir)
    if vault_root is None:
        return {
            "analysisPath": None,
            "analysisVault": None,
        }

    return {
        "analysisPath": note_path.resolve().relative_to(vault_root).as_posix(),
        "analysisVault": vault_root.name,
    }


def _read_analysis_notes(analyses_dir: Path | None) -> list[dict[str, Any]]:
    if analyses_dir is None or not analyses_dir.exists():
        return []

    resolved_analyses_dir = analyses_dir.resolve()
    vault_root = _find_obsidian_vault_root(resolved_analyses_dir)
    notes: list[dict[str, Any]] = []
    for path in sorted(resolved_analyses_dir.glob("*.md")):
        if vault_root is None:
            metadata = {
                "analysisPath": None,
                "analysisVault": None,
            }
        else:
            metadata = {
                "analysisPath": path.resolve().relative_to(vault_root).as_posix(),
                "analysisVault": vault_root.name,
            }
        notes.append(
            {
                "path": str(path),
                "frontmatter": _read_frontmatter(path),
                "body": path.read_text(encoding="utf-8"),
                "analysis_path": metadata["analysisPath"],
                "analysis_vault": metadata["analysisVault"],
            }
        )
    return notes


def _legacy_tag_targets(tag: str) -> tuple[str, str, str] | None:
    lowered = tag.strip().lower()
    if lowered in {"shu", "mir", "das", "agi"}:
        return ("site", lowered.upper(), "occursAt")
    if lowered == "jpt71":
        return ("vessel", "JPT71", "relatedTo")
    if lowered in {"abu_dhabi", "mosb"}:
        return ("hub", "MOSB", "occursAt")
    if lowered == "dsv":
        return ("warehouse", "DSV Indoor", "occursAt")
    return None


def _resolved_operational_anchor(tag: str) -> tuple[str, str, str] | None:
    decision = resolve_location(tag)
    if decision.status == "resolved" and decision.target_id:
        return ("uri", decision.target_id, "occursAt")
    legacy = _legacy_tag_targets(tag)
    if legacy is not None:
        return legacy
    return None


def _read_shipment_rows(excel_path: Path) -> list[dict[str, Any]]:
    rows = _read_excel_rows(excel_path)
    columns = {column for row in rows for column in row.keys()}
    missing = [column for column in REQUIRED_EXCEL_COLUMNS if column not in columns]
    if missing:
        raise ValueError(f"Missing required Excel columns: {', '.join(missing)}")
    return rows


def _read_excel_rows(excel_path: Path, sheet_name: str | None = None) -> list[dict[str, Any]]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name is not None else workbook.active
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = list(next(rows))
    except StopIteration:
        return []

    records: list[dict[str, Any]] = []
    for row in rows:
        record: dict[str, Any] = {}
        for index, header in enumerate(headers):
            if header is None:
                continue
            record[str(header)] = row[index] if index < len(row) else None
        if any(value is not None for value in record.values()):
            records.append(record)
    return records


def _read_jpt_sheets(path: Path) -> dict[str, list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    missing_sheets = [sheet for sheet in REQUIRED_JPT_SHEETS if sheet not in workbook.sheetnames]
    if missing_sheets:
        raise ValueError(f"Missing required JPT sheets: {', '.join(missing_sheets)}")
    return {sheet: _read_excel_rows(path, sheet) for sheet in REQUIRED_JPT_SHEETS}


_ROUTE_FIELDS = (
    ("MOSB", "mosb"),
    ("AGI", "agi"),
    ("DAS", "das"),
    ("MIR", "mir"),
    ("SHU", "shu"),
)


def _clean_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        if value.hour == 0 and value.minute == 0 and value.second == 0 and value.microsecond == 0:
            return value.date().isoformat()
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()

    text = str(value).strip()
    if not text or text.lower() in {"nan", "nat"}:
        return None
    return text


def _shipment_id(shipment_no: str) -> str:
    return f"https://hvdc.logistics/resource/shipment/{shipment_no}"


def _normalize_sources(sources: dict[str, Any]) -> NormalizedSources:
    shipments: list[CanonicalShipment] = []
    cases: list[CanonicalCase] = []
    route_events: list[CanonicalEvent] = []
    document_refs: list[dict[str, Any]] = []
    status_snapshots: list[dict[str, Any]] = []
    charge_candidates: list[dict[str, Any]] = []

    for row in sources.get("shipment_rows", []):
        shipment_no = _clean_text(row.get("SCT SHIP NO."))
        if not shipment_no:
            continue

        shipment_id = _shipment_id(shipment_no)
        shipments.append(
            CanonicalShipment(
                id=shipment_id,
                shipment_no=shipment_no,
                vendor_name=_clean_text(row.get("VENDOR")),
            )
        )

        for location_field, location_slug in _ROUTE_FIELDS:
            event_date = _clean_text(row.get(location_field))
            if not event_date:
                continue
            route_events.append(
                CanonicalEvent(
                    id=(
                        "https://hvdc.logistics/resource/arrival/"
                        f"{shipment_no}/{location_slug}/{event_date}"
                    ),
                    event_type="ArrivalEvent",
                    subject_id=shipment_id,
                    location_id=f"https://hvdc.logistics/resource/site/{location_slug}",
                    event_date=event_date,
                )
            )

        commercial_invoice = _clean_text(row.get("COMMERCIAL INVOICE No."))
        if commercial_invoice:
            document_refs.append(
                {
                    "id": f"https://hvdc.logistics/resource/doc/ci/{commercial_invoice}",
                    "shipment_id": shipment_id,
                    "document_type": "COMMERCIAL_INVOICE",
                    "document_ref": commercial_invoice,
                }
            )

        status_snapshots.append(
            {
                "id": f"https://hvdc.logistics/resource/status/{shipment_no}/current",
                "shipment_id": shipment_id,
                "status": "loaded",
            }
        )

    for row in sources.get("warehouse_rows", []):
        shipment_no = _clean_text(row.get("SCT SHIP NO."))
        case_no = _clean_text(row.get("Case No."))
        if not shipment_no or not case_no:
            continue

        cases.append(
            CanonicalCase(
                id=f"https://hvdc.logistics/resource/case/{shipment_no}/{case_no}",
                shipment_id=_shipment_id(shipment_no),
                case_no=case_no,
            )
        )

    for row in sources.get("inland_cost_rows", []):
        invoice_no = _clean_text(row.get("Invoice No"))
        shipment_no = _clean_text(row.get("Shipment No"))
        if not invoice_no or not shipment_no:
            continue

        charge_candidates.append(
            {
                "invoice_no": invoice_no,
                "shipment_no": shipment_no,
                "row": row,
            }
        )

    return NormalizedSources(
        shipments=shipments,
        cases=cases,
        route_events=route_events,
        document_refs=document_refs,
        status_snapshots=status_snapshots,
        charge_candidates=charge_candidates,
    )


def _load_sources_bundle(
    excel_path: Path,
    warehouse_status_path: Path,
    jpt_reconciled_path: Path,
    inland_cost_path: Path,
    analyses_dir: Path,
) -> LoadedGraphSources:
    source_paths = (
        excel_path,
        warehouse_status_path,
        jpt_reconciled_path,
        inland_cost_path,
    )
    if all(path.exists() for path in source_paths):
        return LoadedGraphSources(
            shipment_rows=_read_shipment_rows(excel_path),
            warehouse_rows=_read_excel_rows(warehouse_status_path),
            jpt_sheets=_read_jpt_sheets(jpt_reconciled_path),
            inland_cost_rows=_read_excel_rows(inland_cost_path),
            analysis_notes=_read_analysis_notes(analyses_dir),
        )

    shipment_rows = _read_shipment_rows(excel_path)
    warehouse_rows = (
        _read_excel_rows(warehouse_status_path) if warehouse_status_path.exists() else []
    )
    inland_cost_rows = _read_excel_rows(inland_cost_path) if inland_cost_path.exists() else []
    return LoadedGraphSources(
        shipment_rows=shipment_rows,
        warehouse_rows=warehouse_rows,
        jpt_sheets={},
        inland_cost_rows=inland_cost_rows,
        analysis_notes=_read_analysis_notes(analyses_dir),
    )


def _resolve_analyses_dir(wiki_dir: Path | None) -> tuple[Path | None, bool]:
    if wiki_dir is not None:
        return wiki_dir, False
    if PRIMARY_ANALYSES_DIR.exists():
        return PRIMARY_ANALYSES_DIR, False
    if DEFAULT_WIKI_DIR.exists():
        return DEFAULT_WIKI_DIR, True
    return None, False


def _build_legacy_shipment_projection(
    shipment_rows: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    nodes: dict[str, dict[str, Any]] = {}
    edges: dict[tuple[str, str, str], dict[str, Any]] = {}

    for row in shipment_rows:
        ship_no = str(row.get("SCT SHIP NO.", "")).strip()
        if not _is_present(ship_no):
            continue

        shipment_id = node_uri("shipment", ship_no)
        nodes[shipment_id] = {
            "data": {
                "id": shipment_id,
                "label": ship_no,
                "type": "Shipment",
                "rdf-schema#label": ship_no,
            }
        }

        po_no = str(row.get("PO No.", "")).strip()
        if _is_present(po_no):
            order_id = node_uri("order", po_no)
            nodes[order_id] = {
                "data": {
                    "id": order_id,
                    "label": po_no,
                    "type": "Order",
                    "rdf-schema#label": po_no,
                }
            }
            edges[(shipment_id, order_id, "hasOrder")] = {
                "data": {
                    "id": f"{shipment_id}|hasOrder|{order_id}",
                    "source": shipment_id,
                    "target": order_id,
                    "label": "hasOrder",
                }
            }

        vendor = str(row.get("VENDOR", "")).strip()
        if _is_present(vendor):
            vendor_id = node_uri("vendor", vendor)
            nodes[vendor_id] = {
                "data": {
                    "id": vendor_id,
                    "label": vendor,
                    "type": "Vendor",
                    "rdf-schema#label": vendor,
                }
            }
            edges[(shipment_id, vendor_id, "suppliedBy")] = {
                "data": {
                    "id": f"{shipment_id}|suppliedBy|{vendor_id}",
                    "source": shipment_id,
                    "target": vendor_id,
                    "label": "suppliedBy",
                }
            }

        vessel = str(row.get("VESSEL NAME/ FLIGHT No.", "")).strip()
        if _is_present(vessel):
            vessel_id = node_uri("vessel", vessel)
            nodes[vessel_id] = {
                "data": {
                    "id": vessel_id,
                    "label": vessel,
                    "type": "Vessel",
                    "rdf-schema#label": vessel,
                }
            }
            edges[(shipment_id, vessel_id, "transportedBy")] = {
                "data": {
                    "id": f"{shipment_id}|transportedBy|{vessel_id}",
                    "source": shipment_id,
                    "target": vessel_id,
                    "label": "transportedBy",
                }
            }

        mosb = str(row.get("MOSB", "")).strip()
        if _is_present(mosb):
            hub_id = node_uri("hub", "MOSB")
            nodes[hub_id] = {
                "data": {
                    "id": hub_id,
                    "label": "MOSB",
                    "type": "Hub",
                    "rdf-schema#label": "MOSB",
                }
            }
            edges[(shipment_id, hub_id, "consolidatedAt")] = {
                "data": {
                    "id": f"{shipment_id}|consolidatedAt|{hub_id}",
                    "source": shipment_id,
                    "target": hub_id,
                    "label": "consolidatedAt",
                }
            }

        for warehouse in WAREHOUSE_COLUMNS:
            if warehouse not in row or not _is_present(row.get(warehouse)):
                continue
            warehouse_id = node_uri("warehouse", warehouse)
            nodes[warehouse_id] = {
                "data": {
                    "id": warehouse_id,
                    "label": warehouse,
                    "type": "Warehouse",
                    "rdf-schema#label": warehouse,
                }
            }
            edges[(shipment_id, warehouse_id, "storedAt")] = {
                "data": {
                    "id": f"{shipment_id}|storedAt|{warehouse_id}",
                    "source": shipment_id,
                    "target": warehouse_id,
                    "label": "storedAt",
                }
            }

        for site in SITE_COLUMNS:
            if site not in row or not _is_present(row.get(site)):
                continue
            site_id = node_uri("site", site)
            nodes[site_id] = {
                "data": {
                    "id": site_id,
                    "label": site,
                    "type": "Site",
                    "rdf-schema#label": site,
                }
            }
            edges[(shipment_id, site_id, "deliveredTo")] = {
                "data": {
                    "id": f"{shipment_id}|deliveredTo|{site_id}",
                    "source": shipment_id,
                    "target": site_id,
                    "label": "deliveredTo",
                }
            }

    return list(nodes.values()), list(edges.values())


def _issue_projection_targets(tag: str) -> tuple[str, str, str] | None:
    target = _resolved_operational_anchor(tag)
    if target is None:
        return None
    if target[0] == "uri":
        return target
    return target


def _build_issue_projection(
    notes: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[str]]:
    nodes: dict[str, dict[str, Any]] = {}
    edges: dict[tuple[str, str, str], dict[str, Any]] = {}
    unresolved: list[str] = []

    for note in notes:
        frontmatter = note.get("frontmatter")
        if not isinstance(frontmatter, dict):
            continue

        title = str(frontmatter.get("title", "")).strip()
        slug = str(frontmatter.get("slug", "")).strip()
        if not title or not slug:
            continue

        issue_id = node_uri("issue", slug)
        nodes[issue_id] = {
            "data": {
                "id": issue_id,
                "label": title,
                "type": "LogisticsIssue",
                "rdf-schema#label": title,
                "analysisPath": note.get("analysis_path"),
                "analysisVault": note.get("analysis_vault"),
            }
        }

        raw_tags = frontmatter.get("tags", [])
        if isinstance(raw_tags, str):
            tags = [raw_tags]
        else:
            tags = [str(tag) for tag in raw_tags]

        projected = False
        for tag in tags:
            target = _issue_projection_targets(tag)
            if not target:
                continue
            kind, label, edge_label = target
            if kind == "uri":
                target_id = label
                target_label = label.rstrip("/").split("/")[-1]
                target_type = "Location"
            else:
                target_id = node_uri(kind, label)
                target_type = kind.capitalize() if kind != "hub" else "Hub"
                target_label = label
            nodes[target_id] = {
                "data": {
                    "id": target_id,
                    "label": target_label,
                    "type": target_type,
                    "rdf-schema#label": target_label,
                }
            }
            edges[(issue_id, target_id, edge_label)] = {
                "data": {
                    "id": f"{issue_id}|{edge_label}|{target_id}",
                    "source": issue_id,
                    "target": target_id,
                    "label": edge_label,
                }
            }
            projected = True

        if not projected:
            unresolved.append(slug)

    return list(nodes.values()), list(edges.values()), unresolved


def _notes_to_canonical_lessons(
    notes: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[str], dict[str, dict[str, str | None]]]:
    lessons: list[dict[str, Any]] = []
    unmapped: list[str] = []
    metadata_by_lesson_id: dict[str, dict[str, str | None]] = {}

    for note in notes:
        frontmatter = note.get("frontmatter")
        if not isinstance(frontmatter, dict):
            continue

        slug = str(frontmatter.get("slug", "")).strip()
        title = str(frontmatter.get("title", "")).strip() or slug
        if not slug:
            continue

        note_kind = resolve_analysis_note(note).get("class_name")
        if note_kind != "IncidentLesson":
            continue

        anchor_id: str | None = None
        raw_tags = frontmatter.get("tags", [])
        tags = [raw_tags] if isinstance(raw_tags, str) else list(raw_tags or [])
        for tag in tags:
            if not isinstance(tag, str):
                continue
            target = _issue_projection_targets(tag)
            if target is None:
                continue
            kind, label, _edge_label = target
            if kind == "uri":
                anchor_id = label
            else:
                anchor_id = node_uri(kind, label)
            break

        if anchor_id is None:
            unmapped.append(slug)
            continue

        lesson_record: dict[str, Any] = {
            "id": node_uri("lesson", slug),
            "label": title,
            "type": "IncidentLesson",
            "analysisPath": note.get("analysis_path"),
            "analysisVault": note.get("analysis_vault"),
        }
        if "/shipment/" in anchor_id:
            lesson_record["shipment_id"] = anchor_id
        elif "/carrier/" in anchor_id:
            lesson_record["carrier_id"] = anchor_id
        elif "/pattern/" in anchor_id:
            lesson_record["pattern_id"] = anchor_id
        else:
            lesson_record["location_id"] = anchor_id
        lessons.append(lesson_record)
        metadata_by_lesson_id[lesson_record["id"]] = {
            "analysisPath": note.get("analysis_path"),
            "analysisVault": note.get("analysis_vault"),
        }

    return lessons, unmapped, metadata_by_lesson_id


def _emit_ttl(graph: Graph, ttl_path: Path) -> None:
    ttl_path.parent.mkdir(parents=True, exist_ok=True)
    graph.serialize(destination=ttl_path, format="turtle")


def _emit_json(nodes: list[dict[str, Any]], edges: list[dict[str, Any]], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)

    (output_dir / "nodes.json").write_text(
        json.dumps(nodes, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (output_dir / "edges.json").write_text(
        json.dumps(edges, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def _emit_audit_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _normalize_for_projection(
    normalized: Any,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    milestone_actual_dt_by_code: dict[tuple[str, str], str] = {}
    for event in getattr(normalized, "milestone_events", []):
        actual_dt = getattr(event, "actual_dt", None)
        if not actual_dt:
            continue
        shipment_id = getattr(event, "shipment_id", None)
        milestone_code = getattr(event, "milestone_code", None)
        if not shipment_id or not milestone_code:
            continue
        milestone_actual_dt_by_code[(shipment_id, milestone_code)] = actual_dt

    shipments = [
        {
            "id": _legacy_dashboard_id(shipment.id),
            "label": shipment.shipment_no,
            "type": "Shipment",
            "country_of_export": getattr(shipment, "country_of_export", None),
            "port_of_loading": getattr(shipment, "port_of_loading", None),
            "port_of_discharge": getattr(shipment, "port_of_discharge", None),
            "ship_mode": getattr(shipment, "ship_mode", None),
            "actual_departure": milestone_actual_dt_by_code.get((shipment.id, "M61")),
            "actual_arrival": milestone_actual_dt_by_code.get((shipment.id, "M80")),
        }
        for shipment in normalized.shipments
    ]
    events = [
        {
            "subject_id": _legacy_dashboard_id(event.subject_id),
            "location_id": _legacy_dashboard_id(event.location_id),
            "location_label": (
                event.location_id.rstrip("/").split("/")[-1].upper() if event.location_id else None
            ),
        }
        for event in normalized.route_events
        if event.location_id
    ]
    return shipments, events, []


def _merge_payloads(*payload_groups: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}
    for payload_group in payload_groups:
        for payload in payload_group:
            data = payload.get("data", {})
            node_id = data.get("id")
            if not node_id:
                continue
            existing = merged.get(node_id)
            if existing is None:
                merged[node_id] = payload
                continue
            existing["data"].update(
                {key: value for key, value in data.items() if value is not None}
            )
    return sorted(
        merged.values(),
        key=lambda item: (
            item["data"]["type"],
            item["data"]["label"],
            item["data"]["id"],
        ),
    )


def _merge_edges(*payload_groups: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: dict[str, dict[str, Any]] = {}
    for payload_group in payload_groups:
        for payload in payload_group:
            data = payload.get("data", {})
            edge_id = data.get("id")
            if not edge_id:
                continue
            merged[edge_id] = payload
    return sorted(
        merged.values(),
        key=lambda item: (
            item["data"]["source"],
            item["data"]["target"],
            item["data"]["label"],
        ),
    )


def export_dashboard_graph_data(
    excel_path: Path = DEFAULT_EXCEL_PATH,
    wiki_dir: Path | None = None,
    output_dir: Path = DEFAULT_OUTPUT_DIR,
    ttl_path: Path | None = DEFAULT_TTL_PATH,
    warehouse_status_path: Path | None = None,
    jpt_reconciled_path: Path | None = None,
    inland_cost_path: Path | None = None,
    audit_dir: Path = DEFAULT_AUDIT_DIR,
) -> None:
    base_dir = excel_path.parent
    warehouse_status_path = warehouse_status_path or base_dir / DEFAULT_WAREHOUSE_STATUS_PATH.name
    jpt_reconciled_path = jpt_reconciled_path or base_dir / DEFAULT_JPT_RECONCILED_PATH.name
    inland_cost_path = inland_cost_path or base_dir / DEFAULT_INLAND_COST_PATH.name
    resolved_wiki_dir, fallback_used = _resolve_analyses_dir(wiki_dir)

    sources = _load_sources_bundle(
        excel_path=excel_path,
        warehouse_status_path=warehouse_status_path,
        jpt_reconciled_path=jpt_reconciled_path,
        inland_cost_path=inland_cost_path,
        analyses_dir=resolved_wiki_dir,
    )
    normalized = normalize_graph_sources(
        {
            "shipment_rows": sources.shipment_rows,
            "warehouse_rows": sources.warehouse_rows,
            "jpt_sheets": sources.jpt_sheets,
            "inland_cost_rows": sources.inland_cost_rows,
            "analysis_notes": sources.analysis_notes,
        }
    )
    knowledge = build_knowledge_objects(sources.analysis_notes)
    canonical_lessons, unmapped_lessons, lesson_metadata_by_id = _notes_to_canonical_lessons(
        sources.analysis_notes
    )
    compatibility_mappings = build_compatibility_mappings()
    dashboard_lessons = [
        {
            **lesson,
            "id": _safe_uri(lesson.get("id")),
            "shipment_id": _legacy_dashboard_id(_safe_uri(lesson.get("shipment_id"))),
            "location_id": _legacy_dashboard_id(_safe_uri(lesson.get("location_id"))),
            "carrier_id": _legacy_dashboard_id(_safe_uri(lesson.get("carrier_id"))),
            "pattern_id": _legacy_dashboard_id(_safe_uri(lesson.get("pattern_id"))),
            "analysisPath": lesson.get("analysisPath"),
            "analysisVault": lesson.get("analysisVault"),
        }
        for lesson in canonical_lessons
    ]

    canonical_graph = build_canonical_graph(
        shipments=[
            {
                "id": _safe_uri(shipment.id),
                "shipment_no": shipment.shipment_no,
                "vendor_name": shipment.vendor_name,
                "country_of_export": shipment.country_of_export,
                "port_of_loading": shipment.port_of_loading,
                "port_of_discharge": shipment.port_of_discharge,
                "ship_mode": shipment.ship_mode,
            }
            for shipment in normalized.shipments
        ],
        cases=[
            {
                "id": _safe_uri(case.id),
                "shipment_id": _safe_uri(case.shipment_id),
                "case_no": case.case_no,
            }
            for case in normalized.cases
        ],
        journey_legs=[
            {
                "id": _safe_uri(leg.id),
                "shipment_id": _safe_uri(leg.shipment_id),
                "origin_port_id": _safe_uri(leg.origin_port_id),
                "origin_port_label": leg.origin_port_label,
                "destination_port_id": _safe_uri(leg.destination_port_id),
                "destination_port_label": leg.destination_port_label,
                "transport_mode": leg.transport_mode,
                "actual_departure": leg.actual_departure,
                "actual_arrival": leg.actual_arrival,
            }
            for leg in getattr(normalized, "journey_legs", [])
        ],
        milestone_events=[
            {
                "id": _safe_uri(event.id),
                "shipment_id": _safe_uri(event.shipment_id),
                "milestone_code": event.milestone_code,
                "actual_dt": event.actual_dt,
                "location_id": _safe_uri(event.location_id),
            }
            for event in getattr(normalized, "milestone_events", [])
        ],
        events=[
            {
                "id": _safe_uri(event.id),
                "event_type": event.event_type,
                "subject_id": _safe_uri(event.subject_id),
                "location_id": _safe_uri(event.location_id),
                "event_date": event.event_date,
            }
            for event in normalized.route_events
        ],
        lessons=[
            {
                **lesson,
                "id": _safe_uri(lesson.get("id")),
                "shipment_id": _safe_uri(lesson.get("shipment_id")),
                "location_id": _safe_uri(lesson.get("location_id")),
                "carrier_id": _safe_uri(lesson.get("carrier_id")),
                "pattern_id": _safe_uri(lesson.get("pattern_id")),
            }
            for lesson in canonical_lessons
        ],
    )
    for triple in compatibility_mappings:
        canonical_graph.add(triple)

    if ttl_path is not None:
        _emit_ttl(canonical_graph, ttl_path)

    dashboard_shipments, dashboard_events, _ = _normalize_for_projection(normalized)
    projected_nodes, projected_edges, projection_audit = build_dashboard_projection(
        shipments=dashboard_shipments,
        events=dashboard_events,
        lessons=dashboard_lessons,
    )
    legacy_nodes, legacy_edges = _build_legacy_shipment_projection(sources.shipment_rows)
    issue_nodes, issue_edges, unresolved_issue_notes = _build_issue_projection(
        sources.analysis_notes
    )

    nodes = _merge_payloads(projected_nodes, legacy_nodes, issue_nodes)
    for node in nodes:
        node_data = node.get("data", {})
        metadata = lesson_metadata_by_id.get(str(node_data.get("id")))
        if metadata:
            node_data.update({key: value for key, value in metadata.items() if value is not None})
    edges = _merge_edges(projected_edges, legacy_edges, issue_edges)
    _emit_json(nodes, edges, output_dir)

    dropped_rows = [
        index
        for index, row in enumerate(sources.shipment_rows)
        if not _is_present(row.get("SCT SHIP NO."))
    ]
    unresolved_attribution = []
    for row in sources.inland_cost_rows:
        if not _is_present(row.get("Shipment No")):
            unresolved_attribution.append(str(row))

    source_audit = {
        "dropped_rows": dropped_rows,
        "loaded_shipments": len(sources.shipment_rows),
        "loaded_notes": len(sources.analysis_notes),
        "selected_analyses_dir": (
            str(resolved_wiki_dir.resolve()) if resolved_wiki_dir is not None else None
        ),
        "analyses_dir_fallback_used": fallback_used,
    }
    resolution_audit = {
        "unresolved_attribution": unresolved_attribution,
        "weak_cost_attribution_confidence": [],
    }
    mapping_audit = {
        "unmapped_lessons": sorted(set(unmapped_lessons + unresolved_issue_notes)),
        "unresolved_compatibility_mappings": [],
        "knowledge_counts": {
            "guides": len(knowledge.guides),
            "rules": len(knowledge.rules),
            "lessons": len(knowledge.lessons),
            "evidence": len(knowledge.evidence),
            "patterns": len(knowledge.patterns),
        },
    }

    _emit_audit_json(audit_dir / "hvdc_ttl_source_audit.json", source_audit)
    _emit_audit_json(audit_dir / "hvdc_ttl_resolution_audit.json", resolution_audit)
    _emit_audit_json(
        audit_dir / "hvdc_ttl_projection_audit.json",
        projection_audit,
    )
    _emit_audit_json(audit_dir / "hvdc_ttl_mapping_audit.json", mapping_audit)


if __name__ == "__main__":
    export_dashboard_graph_data(ttl_path=None)
